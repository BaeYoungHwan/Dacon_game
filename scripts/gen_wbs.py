import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_TITLE='FF4472C4';C_HEADER='FFEEECE1';C_SECTION='FFB6DDE8'
C_ODD='FFFFFFFF';C_EVEN='FFF2F2F2'
C_P0='FFFF0000';C_P1='FFFF6600';C_P2='FF0070C0'
C_DONE='FF00B050';C_WAIT='FF808080'

def fill(h): return PatternFill("solid",fgColor=h)
def mfont(bold=False,color="FF000000",size=9):
    return Font(name="맑은 고딕",bold=bold,color=color,size=size)
def mborder():
    s=Side(style="thin",color="FFB0B0B0")
    return Border(left=s,right=s,top=s,bottom=s)
def malign(h="left",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def ac(ws,row,col,value,bg=None,bold=False,fc="FF000000",wrap=False):
    if bg is None: bg=C_ODD
    c=ws.cell(row=row,column=col,value=value)
    c.fill=fill(bg);c.font=mfont(bold=bold,color=fc,size=9)
    c.border=mborder();c.alignment=malign(wrap=wrap)
    return c

SECTIONS = {
    "1.1": "1.1 분석/기획(AD) | 2026-05-13 | 완료",
    "1.2": "1.2 설계(TD) | 2026-05-13 Day1 오전 | 배영환+신입 병렬",
    "1.3": "1.3 P0-기반 구축 | 2026-05-13 | Hello World+데이터 준비",
    "1.4": "1.4 P1-핵심 기능 개발 | 2026-05-14 | 게임 로직+UI 병렬",
    "1.5": "1.5 P1-통합 연동 | 2026-05-15 | 로직+UI 병합",
    "1.6": "1.6 P2-QA 및 배포 | 2026-05-16 | 게임 완성도+프로덕션",
    "1.7": "1.7 버퍼 | 2026-05-17 | 예상치 못한 이슈 대응",
}

TASKS = [
    ("1.1.1","PRD 작성 v1.1","배영환","완료","05/13","05/13",1,"PRD-v1.md","P0"),
    ("1.1.2","아키텍처 설계","배영환","완료","05/13","05/13",1,"architecture-v1.md, ARD-v1.md","P0"),
    ("1.1.3","WBS 작성 v1","배영환","완료","05/13","05/13",1,"WBS-v1.md","P0"),
    ("1.2.1","[공용] 컴포넌트 Props 명세 작성","배영환","대기","05/13","05/13",0.5,"props-spec.md","P0"),
    ("1.2.2","[Front] 화면 와이어프레임 스케치","신입","대기","05/13","05/13",0.5,"와이어프레임","P1"),
    ("1.2.3","[Back] Supabase 스키마 설계","배영환","대기","05/13","05/13",0.5,"rankings 테이블 DDL","P0"),
    ("1.3.1","[공용] Vite+React 18+Tailwind CSS 초기화","배영환","대기","05/13","05/13",1,"package.json, vite.config.js","P0"),
    ("1.3.2","[공용] 폴더 구조 생성 (pages,components,store,data,lib)","배영환","대기","05/13","05/13",1,"디렉토리 구조","P0"),
    ("1.3.3","[공용] stocks.json 작성 (실제 종목 10개)","신입","대기","05/13","05/13",1,"src/data/stocks.json","P0"),
    ("1.3.4","[공용] news-events.json 작성 (호재/악재 30개)","신입","대기","05/13","05/13",1,"src/data/news-events.json","P0"),
    ("1.3.5","[Back] Supabase 프로젝트+rankings 테이블+RLS","배영환","대기","05/13","05/13",1,"Supabase 프로젝트, .env","P0"),
    ("1.3.6","[배포] Vercel 연결+Hello World 배포 확인","배영환","대기","05/13","05/13",1,"Vercel 프리뷰 URL","P0"),
    ("1.3.7","[Front] 공용 UI 컴포넌트 (Button.jsx, Modal.jsx)","신입","대기","05/13","05/13",1,"src/components/ui/","P1"),
    ("1.4.1","[공용] gameStore 구현 (Zustand-턴,현금,포트폴리오,주가,persist)","배영환","대기","05/14","05/14",1,"src/store/gameStore.js","P0"),
    ("1.4.2","[공용] gameLogic.js-턴 진행+가격 변동 알고리즘","배영환","대기","05/14","05/14",1,"src/lib/gameLogic.js","P0"),
    ("1.4.3","[공용] 매수/매도 로직 구현 (buyStock, sellStock)","배영환","대기","05/14","05/14",1,"gameStore.js 업데이트","P0"),
    ("1.4.4","[공용] 로직 동작 검증-UI 없이 콘솔 테스트","배영환","대기","05/14","05/14",1,"콘솔 출력 확인","P0"),
    ("1.4.5","[Front] StartPage 구현 (닉네임 입력+시작 버튼)","신입","대기","05/14","05/14",1,"src/pages/StartPage.jsx","P0"),
    ("1.4.6","[Front] GamePage 레이아웃 골격 (빈 슬롯 배치)","신입","대기","05/14","05/14",1,"src/pages/GamePage.jsx","P1"),
    ("1.4.7","[Front] StockBoard 컴포넌트 (종목 목록+가격+등락률)","신입","대기","05/14","05/15",1,"src/components/game/StockBoard.jsx","P0"),
    ("1.4.8","[Front] NewsPanel 컴포넌트 (뉴스 헤드라인 출력)","신입","대기","05/14","05/15",1,"src/components/game/NewsPanel.jsx","P1"),
    ("1.4.9","[Front] TurnControl 컴포넌트 (다음 날 버튼+날짜/턴)","신입","대기","05/14","05/15",1,"src/components/game/TurnControl.jsx","P0"),
    ("1.4.10","[Front] Portfolio 컴포넌트 (보유 주식+현금 잔액)","신입","대기","05/14","05/15",1,"src/components/game/Portfolio.jsx","P0"),
    ("1.5.1","[통합] App.jsx 페이지 전환 (Start-Game-Result)","배영환","대기","05/15","05/15",1,"src/App.jsx","P0"),
    ("1.5.2","[통합] GamePage에 gameStore 연결","배영환","대기","05/15","05/15",1,"GamePage.jsx 업데이트","P0"),
    ("1.5.3","[Back] leaderboardStore+Supabase insert/select","배영환","대기","05/15","05/15",1,"src/store/leaderboardStore.js","P1"),
    ("1.5.4","[공용] localStorage persist 연동 (새로고침 복구)","배영환","대기","05/15","05/15",1,"gameStore.js persist 미들웨어","P1"),
    ("1.5.5","[Front] ResultPage 구현 (최종 자산+등급 표시)","신입","대기","05/15","05/15",1,"src/pages/ResultPage.jsx","P0"),
    ("1.5.6","[Front] Leaderboard 컴포넌트 UI (순위 테이블)","신입","대기","05/15","05/15",1,"src/components/leaderboard/Leaderboard.jsx","P1"),
    ("1.5.7","[통합] 통합 버그 수정+PR 리뷰 (신입 Day 2~3)","배영환","대기","05/15","05/15",1,"버그 수정 커밋","P0"),
    ("1.6.1","[QA] 게임 밸런스 플레이테스트 (5~10분)","공통","대기","05/16","05/16",1,"수치 조정 커밋","P0"),
    ("1.6.2","[Front] 조건부 스타일링 (상승 빨강, 하락 파랑, 애니 100~200ms)","신입","대기","05/16","05/16",1,"스타일 업데이트","P1"),
    ("1.6.3","[Front] 반응형 UI 최종 적용 (모바일 375px)","신입","대기","05/16","05/16",1,"반응형 확인","P1"),
    ("1.6.4","[배포] Vercel 환경변수 등록+프로덕션 배포","배영환","대기","05/16","05/16",1,"프로덕션 URL","P0"),
    ("1.6.5","[QA] Lighthouse 성능 점수 확인 (목표: 90+)","배영환","대기","05/16","05/16",1,"Lighthouse 리포트","P2"),
    ("1.6.6","[QA] Supabase RLS 동작 최종 확인","배영환","대기","05/16","05/16",1,"보안 체크리스트","P1"),
    ("1.7.1","[QA] 최종 크로스 브라우저/기기 테스트+잔여 버그","공통","대기","05/17","05/17",1,"최종 확인 체크리스트","P1"),
    ("1.7.2","[배포] 최종 재배포+마감 제출 준비","배영환","대기","05/17","05/17",1,"최종 배포 URL","P0"),
]

SUMMARY=[
    ("1.1 분석/기획",3,3),("1.2 설계",3,0),("1.3 P0 기반",7,0),
    ("1.4 P1 기능",10,0),("1.5 P1 통합",7,0),("1.6 P2 QA",6,0),("1.7 버퍼",2,0),
]
BRANCHES=[
    ("feature/p0-setup","배영환","1.3.1~1.3.6"),("feature/p0-data","신입","1.3.3~1.3.4"),
    ("feature/p0-ui-common","신입","1.3.7"),("feature/p1-game-logic","배영환","1.4.1~1.4.4"),
    ("feature/p1-ui-pages","신입","1.4.5~1.4.6"),("feature/p1-ui-components","신입","1.4.7~1.4.10"),
    ("feature/p1-integration","배영환","1.5.1~1.5.4, 1.5.7"),("feature/p1-ui-result","신입","1.5.5~1.5.6"),
    ("feature/p2-polish","신입","1.6.2~1.6.3"),
]
PC={"P0":C_P0,"P1":C_P1,"P2":C_P2,"P3":C_WAIT}
SC={"완료":C_DONE,"진행중":C_P1,"대기":C_WAIT}
CH=["WBS","태스크","담당자","상태","계획 시작","계획 종료","기간(일)","산출물","우선순위","비고"]
CW=[10,52,10,8,10,10,8,45,10,15]

def build_wbs(ws):
    ws.merge_cells("A1:J1")
    t=ws["A1"]
    t.value="k-stock-merchant  WBS v1.0  |  2026-05-13~2026-05-17  |  배영환(3년차+Claude)+신입"
    t.fill=fill(C_TITLE);t.font=Font(name="맑은 고딕",bold=True,color="FFFFFFFF",size=11)
    t.alignment=malign(h="center");ws.row_dimensions[1].height=22
    ws.merge_cells("A2:J2")
    tt=sum(t2 for _,t2,_ in SUMMARY);td=sum(d for _,_,d in SUMMARY);pct=round(td/tt*100,1)
    sv=("전체 진척도: %d/%d (%.1f%%)  |  " % (td,tt,pct)
        +"  |  ".join("%s %d/%d(%d%%)" % (n,d,t2,round(d/t2*100)) for n,t2,d in SUMMARY))
    sc=ws["A2"];sc.value=sv;sc.fill=fill("FF1F497D")
    sc.font=Font(name="맑은 고딕",color="FFFFFFFF",size=8)
    sc.alignment=Alignment(horizontal="left",vertical="center");ws.row_dimensions[2].height=14
    for ci,(hdr,w) in enumerate(zip(CH,CW),start=1):
        ac(ws,3,ci,hdr,bg=C_HEADER,bold=True)
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=16
    dr=4;ps=None;odd=True
    for task in TASKS:
        sec=".".join(task[0].split(".")[:2])
        if sec!=ps:
            ws.merge_cells(start_row=dr,start_column=1,end_row=dr,end_column=10)
            s=ws.cell(row=dr,column=1,value=SECTIONS[sec])
            s.fill=fill(C_SECTION);s.font=Font(name="맑은 고딕",bold=True,color="FF1F3864",size=9)
            s.border=mborder();s.alignment=malign(h="left");ws.row_dimensions[dr].height=15
            dr+=1;ps=sec;odd=True
        bg=C_ODD if odd else C_EVEN;odd=not odd
        wc,n,o,st,s2,e,d,out,p=task
        ac(ws,dr,1,wc,bg=bg,bold=True);ac(ws,dr,2,n,bg=bg,wrap=True);ac(ws,dr,3,o,bg=bg)
        ac(ws,dr,4,st,bg=bg,fc=SC.get(st,C_WAIT),bold=(st=="완료"))
        ac(ws,dr,5,s2,bg=bg);ac(ws,dr,6,e,bg=bg);ac(ws,dr,7,d,bg=bg)
        ac(ws,dr,8,out,bg=bg,wrap=True);ac(ws,dr,9,p,bg=bg,fc=PC.get(p,C_WAIT),bold=True)
        ac(ws,dr,10,"",bg=bg);ws.row_dimensions[dr].height=15;dr+=1
    dr+=1
    ws.merge_cells(start_row=dr,start_column=1,end_row=dr,end_column=10)
    bh=ws.cell(row=dr,column=1,value="브랜치 전략  (PR: 신입->master | 배영환 리뷰 후 merge)")
    bh.fill=fill("FF4472C4");bh.font=Font(name="맑은 고딕",bold=True,color="FFFFFFFF",size=9)
    bh.border=mborder();bh.alignment=malign(h="left");ws.row_dimensions[dr].height=15;dr+=1
    for ci,h2 in enumerate(["브랜치","담당자","연결 태스크","","","","","","",""],start=1):
        ac(ws,dr,ci,h2,bg=C_HEADER,bold=True)
    ws.row_dimensions[dr].height=14;dr+=1
    for i,(branch,owner,tr) in enumerate(BRANCHES):
        bg=C_ODD if i%2==0 else C_EVEN
        ws.merge_cells(start_row=dr,start_column=3,end_row=dr,end_column=10)
        ac(ws,dr,1,branch,bg=bg,bold=True);ac(ws,dr,2,owner,bg=bg);ac(ws,dr,3,tr,bg=bg)
        for ci in range(4,11):
            c=ws.cell(row=dr,column=ci);c.fill=fill(bg);c.border=mborder()
        ws.row_dimensions[dr].height=14;dr+=1
    ws.freeze_panes="A4"

def build_hist(ws):
    ws.title="History"
    for ci,(h2,w2) in enumerate(zip(["버전","변경사항","작성자","일자"],[10,50,12,14]),start=1):
        ac(ws,1,ci,h2,bg=C_HEADER,bold=True);ws.column_dimensions[get_column_letter(ci)].width=w2
    for ci,val in enumerate(("v1.0","WBS 초안 작성 (38개 태스크, 5일 일정)","배영환","2026-05-13"),start=1):
        ac(ws,2,ci,val,bg=C_ODD,wrap=(ci==2))

wb=Workbook();ws=wb.active;ws.title="WBS";build_wbs(ws)
wh=wb.create_sheet();build_hist(wh)
out=os.path.join("docs","exec-plans","active","WBS-v1.xlsx")
wb.save(out);print("저장 완료:",os.path.abspath(out))
